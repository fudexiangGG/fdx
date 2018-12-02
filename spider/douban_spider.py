import requests as rq
from bs4  import BeautifulSoup as bs
import re,time,random,sys,os
from openpyxl import Workbook
from openpyxl import load_workbook
from fake_useragent import UserAgent

ua =UserAgent()
header={"User-Agent":ua.random}
html=rq.get("https://book.douban.com/tag/?view=type&icn=index-sorttags-all",headers=header)
data=bs(html.text,"html.parser")
length=len(data.find_all(class_='tag-title-wrapper'))
o=0
path='/opt/'
def sheet_conf(sheet):
    sheet['A1'] = "书名"
    sheet['B1'] = "描述"
    sheet['C1'] = "评分"
    sheet['D1'] = "评价人数"
    sheet['E1'] = "简介"


def scan(sheet):
    page = 0
    o = 0
    while o <= num:
        url = 'https://book.douban.com/tag/%s?start=%s&type=T' % (choose, page)
        header = {"User-Agent": ua.random}
        time.sleep(round(random.uniform(0.1, 2.0), 5))
        html = rq.get(url, headers=header)
        data = bs(html.text, "html.parser")
        book_data = data.find_all('li', class_='subject-item')
        if len(book_data) == 0:
            wb.save("%s%s.xlsx" % (path,book_class))
            sys.exit("探索完毕！")

        for i in book_data:
            book = {'book_grade': i.find(class_='rating_nums'), 'book_evaluate': i.find(class_='pl'),
                    'book_name': i.find(class_='info').a, 'book_description': i.find(class_='pub'), 'book_info': i.p}
            for n in book.keys():
                try:
                    if n == 'book_name':
                        book[n] = re.sub('\s', '', book[n].text.strip())
                    elif n == 'book_evaluate':
                        book[n] = re.findall('\d+', book[n].text)[0]
                    elif book[n].text.strip() == '':
                        book[n] = '0'
                    else:
                        book[n] = book[n].text.strip().replace('\n', '')
                except:
                    book[n] =0

            if float(book['book_grade']) >= grade and int(book['book_evaluate']) >= evaluate:
                sheet.append([book['book_name'], book['book_description'], book['book_grade'], book['book_evaluate'],
                book['book_info']])

                o = o + 1
                if o == num:
                    wb.save("%s%s.xlsx" % (path,book_class))
                    sys.exit(0)

        page = page + 20
        print("探索到%s页..." % page)





for i in range(0,length):
    print("\033[0;31m%s\033[0m" % data.find_all(class_='tag-title-wrapper')[i].text.lstrip())
    book_type=data.find_all(class_='article')[0].find('div', class_='').find_all('div', class_='')

    for n in book_type[i].find_all('td'):
        print("%s" % n.text,end="   ")
        o=o+1
        if o== 5:
            print("")
            o = 0

    print("\n")

book_class=input("输入红色子体的大类:")
choose=input("输入你想查看的%s的子类:" % book_class)
grade=float(input("输入书籍评价值(1-10):"))
evaluate=int(input("输入书籍平均人数:"))
num=int(input("书籍的数量:"))


if os.path.exists("%s%s.xlsx" % (path,book_class)):
    wb = load_workbook("%s%s.xlsx" % (path,book_class))
    sheets = wb.sheetnames
    if choose in sheets:
        sheet = wb['%s' % choose]
        scan(sheet)
    else:
        sheet = wb.create_sheet(choose)
        sheet_conf(sheet)
        scan(sheet)
else:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "%s" % choose
    sheet_conf(sheet)
    scan(sheet)






